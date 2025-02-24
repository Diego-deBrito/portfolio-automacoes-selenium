import os
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager

# Caminho do arquivo de entrada e saída
INPUT_FILE = r'C:\Users\diego.brito\OneDrive - Ministério do Desenvolvimento e Assistência Social\Power BI\Mala direta Ofício nova.xlsx'
OUTPUT_FILE = r'C:\Users\diego.brito\OneDrive - Ministério do Desenvolvimento e Assistência Social\Power BI\Python\SEI\retorno.xlsx'

# Lista de documentos a serem verificados
DOCUMENTOS_PROCURADOS = ["Minuta de Termo de Fomento", "Parecer", "Proposta Rejeitada", "Minuta de Convênio"]

def conectar_navegador_existente():
    """Conecta ao navegador Chrome já aberto, utilizando a porta de depuração 9222."""
    try:
        chrome_options = webdriver.ChromeOptions()
        chrome_options.debugger_address = "localhost:9222"
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        print("✅ Conectado ao navegador existente com sucesso.")
        return driver
    except WebDriverException as e:
        print(f"❌ Erro ao conectar ao navegador existente: {e}")
        return None

def clicar_botao_expandir_arvore(driver):
    """Tenta clicar no botão para expandir a árvore de processos dentro do iframe."""
    try:
        driver.switch_to.frame("ifrArvore")
        botao_expandir = WebDriverWait(driver, 2).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "img[title='Abrir todas as Pastas']"))
        )
        botao_expandir.click()
        print("✅ Botão para expandir a árvore clicado com sucesso.")
    except TimeoutException:
        print("⚠️ Botão de expansão não encontrado.")
    finally:
        driver.switch_to.default_content()

def listar_todos_documentos_na_arvore(driver):
    """Lista todos os documentos presentes na árvore de processos."""
    try:
        driver.switch_to.frame("ifrArvore")
        elemento_arvore = WebDriverWait(driver, 2).until(
            EC.presence_of_element_located((By.XPATH, "//form"))
        )
        documentos_elementos = elemento_arvore.find_elements(By.XPATH, ".//a | .//div")
        documentos = [elemento.text.strip() for elemento in documentos_elementos if elemento.text.strip()]
        print(f"📄 Documentos capturados: {documentos}")
        return documentos
    except TimeoutException:
        print("⚠️ Árvore de processos não carregada.")
        return []
    finally:
        driver.switch_to.default_content()

def buscar_processo(driver, numero_processo):
    """Busca o número do processo e verifica a presença dos documentos especificados."""
    try:
        campo_busca = WebDriverWait(driver, 2).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/div[1]/div[4]/span/form/input"))
        )
        campo_busca.clear()
        campo_busca.send_keys(numero_processo + Keys.RETURN)
        print(f"🔍 Buscando processo '{numero_processo}'.")

        time.sleep(2)  # Reduzindo o tempo de espera para 2 segundos

        clicar_botao_expandir_arvore(driver)
        documentos = listar_todos_documentos_na_arvore(driver)

        resultados = {doc: "SIM" if any(doc.lower() in d.lower() for d in documentos) else "NÃO" for doc in DOCUMENTOS_PROCURADOS}
        return resultados
    except Exception as e:
        print(f"❌ Erro ao buscar processo '{numero_processo}': {e}")
        return {doc: "Erro" for doc in DOCUMENTOS_PROCURADOS}

def salvar_resultado(numero_processo, resultados, workbook, sheet):
    """Salva os resultados na planilha de saída."""
    linha = [numero_processo] + [resultados.get(doc, "Erro") for doc in DOCUMENTOS_PROCURADOS]
    sheet.append(linha)
    print(f"✅ Resultado salvo para o processo '{numero_processo}': {linha}")

def main():
    driver = conectar_navegador_existente()
    if not driver:
        return

    # Carrega a planilha uma única vez
    dataframe = pd.read_excel(INPUT_FILE, sheet_name="Planilha1")
    dataframe.columns = dataframe.columns.str.strip().str.lower()

    if not os.path.exists(OUTPUT_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Resultados"
        sheet.append(["Processo"] + DOCUMENTOS_PROCURADOS)
    else:
        workbook = load_workbook(OUTPUT_FILE)
        sheet = workbook.active

    for _, linha in dataframe.iterrows():
        numero_processo = str(linha["processo"]).strip()
        resultados = buscar_processo(driver, numero_processo)
        salvar_resultado(numero_processo, resultados, workbook, sheet)

        # Salva a planilha após cada iteração
        workbook.save(OUTPUT_FILE)

    driver.quit()
    print("🏁 Execução concluída!")

if __name__ == "__main__":
    main()

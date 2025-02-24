import os
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager

# Caminho do arquivo de entrada e saída
INPUT_FILE = r'C:\path\to\input.xlsx'
OUTPUT_FILE = r'C:\path\to\resultados.xlsx'

def conectar_navegador():
    """Inicializa o navegador Chrome usando WebDriverManager."""
    try:
        print("Iniciando navegador...")
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        driver.maximize_window()
        print("✅ Navegador iniciado com sucesso.")
        return driver
    except WebDriverException as exception:
        print(f"❌ Erro ao iniciar o navegador: {exception}")
        return None

def buscar_email_por_cnpj(driver, cnpj):
    """Busca o e-mail associado ao CNPJ em um site específico."""
    try:
        # Substitua pela URL do site em que deseja buscar o CNPJ
        URL_BUSCA = 'https://www.exemplo.com/busca-cnpj'
        driver.get(URL_BUSCA)

        # 1 - Acessar o primeiro elemento na página
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]"))
        )

        # 2 - Clicar no link específico para prosseguir
        link_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[3]/div[2]/div[1]/div[1]/ul/li[1]/a"))
        )
        link_element.click()
        time.sleep(2)  # Pequena pausa para carregar a página

        # 3 - Localizar o campo de busca de CNPJ e inserir o valor
        campo_cnpj = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div/div[2]/div/div/div[3]/div[1]/div/form/fieldset/div[1]/div[1]/input"))
        )
        campo_cnpj.clear()
        campo_cnpj.send_keys(cnpj)

        # 4 - Clicar no botão de consultar
        botao_consultar = driver.find_element(By.XPATH, "/html/body/div[1]/div/div[2]/div/div/div[3]/div[1]/div/form/fieldset/div[3]/div[2]/input")
        botao_consultar.click()
        print(f"🔍 Buscando e-mail para o CNPJ: {cnpj}")

        # 5 - Aguardar o link de resultado aparecer e clicar nele
        resultado_link = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div/div[2]/div/div/div[3]/div[1]/div/div/div/div[2]/table/tbody/tr/td/h3/a"))
        )
        resultado_link.click()
        time.sleep(2)  # Pausa para carregar a página com os detalhes

        # 6 - Capturar o e-mail associado ao CNPJ
        email_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div/div[2]/div/div/div[3]/div[2]/div[2]/div[1]/div/div[1]/div[10]"))
        )
        email = email_element.text.strip()
        print(f"✅ E-mail encontrado: {email}")
        return email

    except TimeoutException:
        print(f"❌ Tempo esgotado ao buscar o CNPJ: {cnpj}")
        return "Email não encontrado"
    except NoSuchElementException:
        print(f"❌ E-mail não encontrado para o CNPJ: {cnpj}")
        return "Email não encontrado"
    except Exception as e:
        print(f"❌ Erro inesperado: {e}")
        return "Erro na busca"

def salvar_resultado(cnpj, email):
    """Salva os resultados na planilha de saída."""
    try:
        if not os.path.exists(OUTPUT_FILE):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Resultados"
            sheet.append(["CNPJ", "Email"])
        else:
            workbook = load_workbook(OUTPUT_FILE)
            sheet = workbook.active

        linha = [cnpj, email]
        sheet.append(linha)
        workbook.save(OUTPUT_FILE)
        print(f"✅ Resultado salvo para o CNPJ '{cnpj}': {email}")
    except Exception as exception:
        print(f"❌ Erro ao salvar o resultado no Excel: {exception}")

def main():
    driver = conectar_navegador()
    if not driver:
        return

    # Lê a planilha de entrada
    dataframe = pd.read_excel(INPUT_FILE)

    # Certifique-se de que a coluna 'cnpj' existe
    if 'cnpj' not in dataframe.columns.str.lower():
        print("❌ Coluna 'cnpj' não encontrada na planilha de entrada.")
        driver.quit()
        return

    # Loop pelos CNPJs e busca os e-mails
    for _, linha in dataframe.iterrows():
        cnpj = str(linha['cnpj']).strip()
        email = buscar_email_por_cnpj(driver, cnpj)
        salvar_resultado(cnpj, email)

    driver.quit()
    print("🚀 Processo concluído!")

if __name__ == "__main__":
    main()

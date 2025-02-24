import os
import pyperclip
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from openpyxl import Workbook, load_workbook

def conectar_navegador_existente():
    """ Conecta ao navegador Chrome já aberto na porta de depuração 9222. """
    try:
        print("🚀 Tentando conectar ao navegador na porta 9222...")
        opcoes_navegador = webdriver.ChromeOptions()
        opcoes_navegador.debugger_address = "localhost:9222"
        navegador = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opcoes_navegador)
        print("✅ Conectado ao navegador existente com sucesso.")
        return navegador
    except WebDriverException as erro:
        print(f"❌ Erro ao conectar ao navegador existente: {erro}")
        return None

def clicar_rapido(navegador, xpath, tempo_espera=5):
    """ Usa `execute_script` para clicar mais rápido, evitando esperas desnecessárias. """
    try:
        elemento = WebDriverWait(navegador, tempo_espera).until(EC.presence_of_element_located((By.XPATH, xpath)))
        navegador.execute_script("arguments[0].click();", elemento)  # CLIQUE MAIS RÁPIDO
        print(f"✔️ Elemento clicado rapidamente: {xpath}")
    except (TimeoutException, NoSuchElementException) as erro:
        print(f"⚠️ Erro ao clicar no elemento {xpath}: {erro}")

def colar_texto(navegador, xpath):
    """ Cola o conteúdo da área de transferência de forma mais eficiente. """
    try:
        elemento = WebDriverWait(navegador, 3).until(EC.element_to_be_clickable((By.XPATH, xpath)))
        elemento.click()
        elemento.send_keys(Keys.CONTROL, 'v')  # Cola o texto copiado
        time.sleep(0.5)  # Reduz tempo de espera
        print(f"📋 Texto colado no campo: {xpath}")
    except (TimeoutException, NoSuchElementException) as erro:
        print(f"⚠️ Erro ao colar texto no elemento {xpath}: {erro}")

def criar_ou_atualizar_planilha(caminho_arquivo, linha_dados):
    """ Cria ou adiciona uma linha ao arquivo Excel de forma otimizada. """
    colunas_saida = ["Instrumento nº", "Técnico", "e-mail do Técnico", "AjustesPT", "Data da Solicitação"]

    if os.path.exists(caminho_arquivo):
        planilha = load_workbook(caminho_arquivo)
        aba = planilha.active
    else:
        planilha = Workbook()
        aba = planilha.active
        aba.append(colunas_saida)  # Adiciona cabeçalhos

    aba.append(linha_dados)  # Adiciona a linha processada
    planilha.save(caminho_arquivo)
    print(f"✅ Dados salvos para o instrumento {linha_dados[0]}")

def executar_processo_principal():
    """ Fluxo principal para carregar dados do Excel, processar informações e salvar em outra planilha. """
    print("🚀 Iniciando o processo principal...")

    navegador = conectar_navegador_existente()
    if not navegador:
        print("❌ Não foi possível conectar ao navegador. Encerrando o processo.")
        return

    caminho_arquivo_entrada = r'C:/Users/diego.brito/Downloads/robov1/CONTROLE DE PARCERIAS CGAP.xlsx'
    caminho_arquivo_saida = r'C:/Users/diego.brito/Downloads/robov1/Resultados_Ajuste.xlsx'

    try:
        dataframe = pd.read_excel(caminho_arquivo_entrada, sheet_name='PARCERIAS CGAP', engine='openpyxl')

        colunas_interesse = ["Instrumento nº", "Técnico", "e-mail do Técnico"]
        if not all(coluna in dataframe.columns for coluna in colunas_interesse):
            print(f"❌ Colunas necessárias não encontradas: {colunas_interesse}. Encerrando o processo.")
            return

        dataframe["Instrumento nº"] = dataframe["Instrumento nº"].astype(str).str.replace(r'\.0$', '', regex=True)

        for _, linha in dataframe.iterrows():
            instrumento_numero = linha["Instrumento nº"]
            tecnico = linha["Técnico"]
            email_tecnico = linha["e-mail do Técnico"]

            pyperclip.copy(instrumento_numero)

            try:
                # Fluxo de navegação otimizado
                clicar_rapido(navegador, '//*[@id="menuPrincipal"]/div[1]/div[4]')
                clicar_rapido(navegador, '/html/body/div[1]/div[3]/div[2]/div[1]/div[1]/ul/li[6]/a')
                colar_texto(navegador, '//*[@id="consultarNumeroConvenio"]')
                clicar_rapido(navegador, '//*[@id="form_submit"]')
                clicar_rapido(navegador, '//*[@id="instrumentoId"]/a')
                clicar_rapido(navegador, '//*[@id="div_-173460853"]/span/span')
                clicar_rapido(navegador, '//*[@id="menu_link_-173460853_-1293190284"]/div/span/span')
                clicar_rapido(navegador, '/html/body/div[3]/div[2]/div[1]/a')

                # Captura de situação otimizada
                try:
                    elemento_situacao = WebDriverWait(navegador, 3).until(
                        EC.presence_of_element_located((By.XPATH, '//*[@id="row"]//td[contains(text(),"Em Análise")]'))
                    )
                    situacao = elemento_situacao.text
                    clicar_rapido(navegador, '//*[@id="tbodyrow"]/tr[5]/td[4]/nobr/a')
                    data_solicitacao = navegador.find_element(By.XPATH, '//*[@id="tr-editarDataSolicitacao"]/td[2]').text
                except TimeoutException:
                    situacao = "Sem ajuste"
                    data_solicitacao = ""

                clicar_rapido(navegador, '//*[@id="logo"]/a/span')

                # Salva os dados imediatamente após processar
                criar_ou_atualizar_planilha(caminho_arquivo_saida, [instrumento_numero, tecnico, email_tecnico, situacao, data_solicitacao])

            except Exception as erro:
                print(f"❌ Erro ao processar o instrumento {instrumento_numero}: {erro}")
                continue

    except Exception as erro:
        print(f"❌ Erro ao carregar ou processar o arquivo de entrada: {erro}")

    finally:
        navegador.quit()
        print("✅ Processo concluído.")

if __name__ == "__main__":
    executar_processo_principal()

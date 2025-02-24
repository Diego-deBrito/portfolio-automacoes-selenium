import os
import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime


def conectar_navegador_existente(retentativas=3):
    """
    Conecta ao navegador Chrome já aberto, utilizando a porta de depuração 9222, com múltiplas tentativas.
    """
    for tentativa in range(1, retentativas + 1):
        try:
            print(f"[INFO] Tentativa {tentativa} de conectar ao navegador na porta 9222...")
            options = webdriver.ChromeOptions()
            options.debugger_address = "localhost:9222"
            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
            print("[INFO] Conectado ao navegador existente com sucesso.")
            return driver
        except WebDriverException as e:
            print(f"[ERRO] Erro ao conectar ao navegador (tentativa {tentativa}): {e}")
            time.sleep(3)
    print("[ERRO] Não foi possível conectar ao navegador após múltiplas tentativas.")
    return None


def inicializar_planilha(caminho_arquivo_excel):
    """
    Inicializa o arquivo Excel de saída e cria o cabeçalho se ele ainda não existir.
    """
    if not os.path.exists(caminho_arquivo_excel):
        workbook = openpyxl.Workbook()
        planilha = workbook.active
        planilha.title = "Dados Coletados"
        planilha.append(["CNPJ", "Descrição do Objeto"])  # Cabeçalho
        workbook.save(caminho_arquivo_excel)
        print(f"[INFO] Planilha '{caminho_arquivo_excel}' criada com sucesso.")
    else:
        print(f"[INFO] Planilha '{caminho_arquivo_excel}' já existe. Dados serão adicionados.")


def salvar_dados(caminho_arquivo_excel, cnpj, descricao):
    """
    Salva o CNPJ e a Descrição do Objeto no arquivo Excel.
    """
    try:
        workbook = openpyxl.load_workbook(caminho_arquivo_excel)
        planilha = workbook.active
        planilha.append([cnpj, descricao])
        workbook.save(caminho_arquivo_excel)
        print(f"[INFO] Dados salvos com sucesso: CNPJ={cnpj}, Descrição={descricao}")
    except Exception as e:
        print(f"[ERRO] Falha ao salvar os dados na planilha: {e}")


def ler_cnpjs_da_planilha(caminho_arquivo_excel, aba_nome):
    """
    Lê a coluna B chamada 'CNPJ' de uma planilha Excel.
    """
    try:
        if not os.path.exists(caminho_arquivo_excel):
            print(f"[ERRO] Planilha de entrada '{caminho_arquivo_excel}' não encontrada.")
            return []

        workbook = openpyxl.load_workbook(caminho_arquivo_excel)
        aba = workbook[aba_nome]
        cnpjs = []

        # Localiza a coluna B com cabeçalho 'CNPJ' e extrai os valores
        for linha in aba.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):  # Apenas coluna B
            cnpj = linha[0]  # Coluna B (index 0 para iter_rows)
            if cnpj:  # Garante que o valor não é vazio
                cnpjs.append(str(cnpj))

        print(f"[INFO] Total de {len(cnpjs)} CNPJs carregados da planilha.")
        return cnpjs
    except Exception as e:
        print(f"[ERRO] Falha ao ler CNPJs da planilha: {e}")
        return []


def executar_fluxo_novo(driver, cnpj, caminho_arquivo_excel):
    """
    Executa o novo fluxo:
    1. Clica no menu principal.
    2. Navega para o submenu.
    3. Pesquisa pelo CNPJ.
    4. Extrai a "Descrição do Objeto" e salva no Excel.
    """
    try:
        print("[INFO] Iniciando novo fluxo para o CNPJ:", cnpj)

        # Passo 1: Clica no menu principal
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]'))).click()
        print("[INFO] Passo 1: Menu principal clicado.")

        # Passo 2: Clica no submenu
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[3]/div[2]/div[1]/div[1]/ul/li[1]/a'))).click()
        print("[INFO] Passo 2: Submenu clicado.")

        # Passo 3: Preenche o campo de pesquisa com o CNPJ
        campo_cnpj = WebDriverWait(driver, 10).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[1]/div/div[2]/div/div/div[3]/div[1]/div/form/fieldset/div[1]/div[1]/input')))
        campo_cnpj.clear()
        campo_cnpj.send_keys(cnpj)
        print("[INFO] Passo 3: CNPJ inserido no campo de pesquisa.")

        # Passo 4: Clica no botão "Consultar"
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,
                                                                    '/html/body/div[1]/div/div[2]/div/div/div[3]/div[1]/div/form/fieldset/div[3]/div[2]/input'))).click()
        print("[INFO] Passo 4: Botão 'Consultar' clicado.")

        # Passo 5: Clica no resultado da pesquisa
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,
                                                                    '/html/body/div[1]/div/div[2]/div/div/div[3]/div[1]/div/div/div/div[2]/table/tbody/tr/td/h3/a'))).click()
        print("[INFO] Passo 5: Resultado da pesquisa clicado.")

        # Passo 6: Navega para a aba "Descrição do Objeto"
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
            (By.XPATH, '/html/body/div[1]/div/div[2]/div/div/div[3]/div[2]/div[2]/div[1]/ul/li[2]/a'))).click()
        print("[INFO] Passo 6: Aba 'Descrição do Objeto' clicada.")
        # Passo 7: Extrai a "Descrição do Objeto"
        descricao_objeto = WebDriverWait(driver, 10).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[1]/div/div[2]/div/div/div[3]/div[2]/div[2]/div[1]/div'))).text
        print(f"[INFO] Passo 7: Descrição do Objeto extraída: {descricao_objeto}")

        # Salva os dados no Excel
        salvar_dados(caminho_arquivo_excel, cnpj, descricao_objeto)

    except TimeoutException as e:
        print(f"[ERRO] Elemento não encontrado no fluxo do CNPJ {cnpj}: {e}")
    except Exception as e:
        print(f"[ERRO] Erro inesperado no fluxo do CNPJ {cnpj}: {e}")


def executar_processo_novo():
    """
    Função principal que executa o processo para os CNPJs.
    """
    # Caminho para salvar os dados coletados
    caminho_arquivo_excel_saida = r"C:\Users\diego.brito\Downloads\robov1\descricao_objeto.xlsx"

    # Caminho da planilha de entrada
    caminho_arquivo_excel_entrada = r"C:\Users\diego.brito\OneDrive - Ministério do Desenvolvimento e Assistência Social\Power BI\Lista - Entidades.xlsx"
    aba_nome = "Relação"  # Nome da aba onde os CNPJs estão localizados

    # Inicializa a planilha de saída
    inicializar_planilha(caminho_arquivo_excel_saida)
    # Lê os CNPJs da planilha de entrada
    lista_cnpjs = ler_cnpjs_da_planilha(caminho_arquivo_excel_entrada, aba_nome)
    if not lista_cnpjs:
        print("[ERRO] Nenhum CNPJ encontrado na planilha. Processo encerrado.")
        return
    # Conecta ao navegador já existent
    driver = conectar_navegador_existente()
    if not driver:
        print("[ERRO] Não foi possível conectar ao navegador. Processo encerrado.")
        return
    try:
        # Processa cada CNPJ
        for cnpj in lista_cnpjs:
            executar_fluxo_novo(driver, cnpj, caminho_arquivo_excel_saida)
    finally:
        driver.quit()
        print("[INFO] Processo concluído e navegador fechado.")


if __name__ == "__main__":
    executar_processo_novo()
